import openpyxl
def create_excel_sheet(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "student Detials"
    sheet["A1"] = "student Name"
    sheet["B1"] = "student Id"
    sheet["C1"] = "Gender"
    sheet["D1"] = "Present days"
    sheet["E1"] = "Total days"
    workbook.save(file_name)
def append_to_excel(file_name, name, Id , gender,Presentdays,Totaldays):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook["student Detials"]  
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=name)
    sheet.cell(row=next_row, column=2, value=Id)
    sheet.cell(row=next_row, column=3, value=gender)
    sheet.cell(row=next_row, column=4, value=Presentdays)
    sheet.cell(row=next_row, column=5, value=Totaldays)
    workbook.save(file_name)
def main():
    file_name = "student.xlsx"
    try:
        create_excel_sheet(file_name)
    except Exception as e:
        print("Error creating Excel sheet:", e)
        return
    while True:
        name = input("Enter student Name: (or 'exit' to stop): ")
        if name.lower() == "exit":
            break
    
        Id=int(input("Enter Id:"))
        age =int(input("Enter age: "))
        gender=input("Enter Gender:")
        Presentdays=int(input("Enter present days:"))
        Totaldays=int(input("Enter total days:"))
        try:
            append_to_excel(file_name,name ,Id, gender,Presentdays,Totaldays)
            print("Successfully added data to the Excel sheet.")
        except Exception as e:
            print("Error appending data to Excel sheet:", e)
if __name__ == "__main__":
    main()